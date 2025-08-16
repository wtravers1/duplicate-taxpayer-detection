"""
Delinquent Account Report Automation Tool

Purpose:
    Automates login to a county tax system and retrieves delinquent Real Estate (RES)
    and Personal Property (VPP) account detail reports.

    Identifies and exports:
        - Customers with both RES and VPP accounts
        - Probable matches between RES and VPP customers using fuzzy matching
        - VPP customers with multiple active accounts

Requirements:
    - Python 3.12+
    - pandas
    - rapidfuzz
    - openpyxl
    - Custom automation library (tech_library)
    - Windows environment with appropriate browser drivers
"""

import os
import sys
import pandas as pd
from rapidfuzz import process, fuzz
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Border, Side
from datetime import datetime

# Import custom automation library
# Note: Replace with your actual automation library path
sys.path.append(r"path/to/your/automation/library")
import tech_library as tech

# Configuration Constants
FUZZY_MATCH_THRESHOLD = 85
STREET_SIMILARITY_HIGHLIGHT_THRESHOLD = 80
OUTPUT_FOLDER = r"path/to/output/folder"


def normalize_customer_keys(df):
    """Normalize customer keys by removing commas and converting to string."""
    df["Customer Key"] = df["Customer Key"].astype(str).str.replace(",", "")
    return df


def create_broad_summary(res_df, vpp_df):
    """
    Create summary of customers with both RES and VPP accounts.

    Args:
        res_df (pd.DataFrame): Real Estate accounts dataframe
        vpp_df (pd.DataFrame): Personal Property accounts dataframe

    Returns:
        pd.DataFrame: Summary of combined accounts
    """
    # Find common customer keys
    common_keys = set(res_df["Customer Key"]) & set(vpp_df["Customer Key"])

    # Filter dataframes to common keys
    res_trim = res_df[res_df["Customer Key"].isin(common_keys)].copy()
    vpp_trim = vpp_df[vpp_df["Customer Key"].isin(common_keys)].copy()

    # Combine and summarize
    combined = pd.concat([res_trim, vpp_trim])

    summary_df = (
        combined.groupby(["Customer Key", "Customer Name"])
        .agg({"Account ID": lambda x: ", ".join(sorted(x)), "Total Balance": "sum"})
        .sort_values(by="Total Balance", ascending=False)
        .reset_index()
    )

    # Clean up data
    summary_df = summary_df[summary_df["Customer Key"] != "Totals"]
    summary_df["Customer Key"] = r"\c" + summary_df["Customer Key"].astype(str)

    return summary_df


def perform_fuzzy_matching(res_df, vpp_df, threshold=FUZZY_MATCH_THRESHOLD):
    """
    Find probable matches between RES and VPP customers using fuzzy name matching.

    Args:
        res_df (pd.DataFrame): Real Estate accounts dataframe
        vpp_df (pd.DataFrame): Personal Property accounts dataframe
        threshold (int): Minimum similarity score for matches

    Returns:
        pd.DataFrame: Probable matches with similarity scores
    """
    # Get customers that only appear in one dataset
    res_only = res_df[~res_df["Customer Key"].isin(vpp_df["Customer Key"])].copy()
    vpp_only = vpp_df[~vpp_df["Customer Key"].isin(res_df["Customer Key"])].copy()

    # Clean data
    res_only = res_only.dropna(subset=["Customer Name"]).drop_duplicates(
        subset=["Customer Key", "Customer Name"]
    )
    vpp_only = vpp_only.dropna(subset=["Customer Name"]).drop_duplicates(
        subset=["Customer Key", "Customer Name"]
    )

    matches = []
    vpp_names = list(vpp_only["Customer Name"])

    for _, res_row in res_only.iterrows():
        res_name = res_row["Customer Name"]
        res_key = res_row["Customer Key"]

        match = process.extractOne(
            res_name, vpp_names, scorer=fuzz.token_sort_ratio, score_cutoff=threshold
        )

        if match:
            matched_name, score, match_index = match
            vpp_match_row = vpp_only.iloc[match_index]

            if vpp_match_row["Customer Key"] != res_key:
                # Calculate street similarity if street data is available
                res_street = res_row.get("Street", "")
                vpp_street = vpp_match_row.get("Street", "")
                street_score = (
                    fuzz.token_sort_ratio(res_street, vpp_street)
                    if all(isinstance(x, str) for x in [res_street, vpp_street])
                    else None
                )

                matches.append(
                    {
                        "RES Customer Key": r"\c" + res_key,
                        "RES Name": res_name,
                        "RES Street": res_street,
                        "VPP Customer Key": r"\c" + vpp_match_row["Customer Key"],
                        "VPP Name": matched_name,
                        "VPP Street": vpp_street,
                        "Name Similarity": score,
                        "Street Similarity": street_score,
                    }
                )

    matches_df = pd.DataFrame(matches).sort_values(
        by="Street Similarity", ascending=False
    )

    return matches_df


def find_multi_account_customers(vpp_df):
    """
    Find VPP customers with multiple active accounts.

    Args:
        vpp_df (pd.DataFrame): Personal Property accounts dataframe

    Returns:
        pd.DataFrame: Customers with multiple accounts
    """
    duplicate_keys = vpp_df["Customer Key"].value_counts()[lambda x: x > 1].index
    duplicate_df = vpp_df[vpp_df["Customer Key"].isin(duplicate_keys)].copy()
    duplicate_df = duplicate_df.sort_values(by=["Customer Key", "Account ID"])
    duplicate_df["Customer Key"] = r"\c" + duplicate_df["Customer Key"].astype(str)

    return duplicate_df


def auto_adjust_column_width(ws):
    """Auto-adjust column widths based on content."""
    for col in ws.columns:
        max_length = 0
        column = col[0].column
        column_letter = get_column_letter(column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except (AttributeError, TypeError) as e:
                print(
                    f"Warning: Could not read cell value in column {column_letter}: {e}"
                )
        adjusted_width = max_length + 2
        ws.column_dimensions[column_letter].width = adjusted_width


def format_header(ws):
    """Format header row with bold font and borders."""
    bold_font = Font(bold=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for cell in ws[1]:
        cell.font = bold_font
        cell.border = thin_border


def highlight_high_similarity_matches(ws_matches):
    """Highlight fuzzy matches with high street similarity."""
    try:
        sim_col = [cell.value for cell in ws_matches[1]].index("Street Similarity") + 1
        highlight_fill = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid"
        )

        for row in ws_matches.iter_rows(min_row=2):
            cell = row[sim_col - 1]
            try:
                if float(cell.value) > STREET_SIMILARITY_HIGHLIGHT_THRESHOLD:
                    for c in row:
                        c.fill = highlight_fill
            except (ValueError, TypeError):
                continue
    except ValueError:
        pass  # Column not found


def apply_alternating_colors(ws_multi):
    """Apply alternating colors to multi-account customers grouped by Customer Key."""
    grey_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    white_fill = PatternFill(
        start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
    )

    try:
        custkey_col = [cell.value for cell in ws_multi[1]].index("Customer Key") + 1
        prev_key = None
        use_grey = True

        for row in ws_multi.iter_rows(min_row=2):
            current_key = row[custkey_col - 1].value
            if current_key != prev_key:
                use_grey = not use_grey
                prev_key = current_key
            row_fill = grey_fill if use_grey else white_fill
            for cell in row:
                cell.fill = row_fill
    except ValueError:
        pass  # Column not found


def export_to_excel(summary_df, matches_df, duplicate_df, output_folder=OUTPUT_FOLDER):
    """
    Export all dataframes to a single Excel workbook with multiple sheets.

    Args:
        summary_df (pd.DataFrame): Summary of combined accounts
        matches_df (pd.DataFrame): Fuzzy matches
        duplicate_df (pd.DataFrame): Multi-account customers
        output_folder (str): Path to output folder

    Returns:
        str: Path to created file
    """
    # Create output folder if it doesn't exist
    os.makedirs(output_folder, exist_ok=True)

    file_path = os.path.join(
        output_folder, f"RES_VPP_Comparison_{datetime.now().strftime('%Y%m%d')}.xlsx"
    )

    wb = Workbook()

    # Summary Sheet
    ws_summary = wb.active
    ws_summary.title = "Combined Accounts Summary"
    for r in dataframe_to_rows(summary_df, index=False, header=True):
        ws_summary.append(r)
    format_header(ws_summary)
    auto_adjust_column_width(ws_summary)

    # Fuzzy Matches Sheet
    ws_matches = wb.create_sheet("Probable Matches")
    for r in dataframe_to_rows(matches_df, index=False, header=True):
        ws_matches.append(r)
    format_header(ws_matches)
    auto_adjust_column_width(ws_matches)
    highlight_high_similarity_matches(ws_matches)

    # Multi-Account Sheet
    ws_multi = wb.create_sheet("Multi-Account Customers")
    for r in dataframe_to_rows(duplicate_df, index=False, header=True):
        ws_multi.append(r)
    format_header(ws_multi)
    auto_adjust_column_width(ws_multi)
    apply_alternating_colors(ws_multi)

    wb.save(file_path)
    return file_path


def main():
    """Main execution function."""
    print("Starting delinquent account report automation...")

    # Configuration for report parameters
    res_report_args = {
        '//input[@id="report-type-selector"]': "Real Estate",
        '//input[@id="status-selector"]': "1",  # Delinquent status
    }
    vpp_report_args = {
        '//input[@name="report-type-selector"]': "Personal Property",
        '//input[@name="status-selector"]': "1",  # Delinquent status
    }

    # Get current user
    username = os.getlogin()
    print(f"Logged in as: {username}")

    # Initialize web driver
    driver = tech.create_driver()

    try:
        # Login to system
        tech.open_system(driver)
        tech.system_login(driver, username)
        print("Successfully logged into system")

        # Generate Real Estate report
        print("Generating Real Estate delinquent report...")
        tech.choose_report(
            driver, report_name="Delinquent Account Detail", parameters=res_report_args
        )
        download = tech.choose_view(driver, view_name="detailed_view")
        res_df = tech.process_download(downloads_list=download)
        tech.close_report_manager(driver)

        # Generate Personal Property report
        print("Generating Personal Property delinquent report...")
        tech.choose_report(
            driver, report_name="Delinquent Account Detail", parameters=vpp_report_args
        )
        download = tech.choose_view(driver, view_name="detailed_view")
        vpp_df = tech.process_download(downloads_list=download)

        # Data processing
        print("Processing data...")
        res_df = normalize_customer_keys(res_df)
        vpp_df = normalize_customer_keys(vpp_df)

        # Analysis
        print("Creating broad summary...")
        summary_df = create_broad_summary(res_df, vpp_df)

        print("Performing fuzzy matching...")
        matches_df = perform_fuzzy_matching(res_df, vpp_df)

        print("Finding multi-account customers...")
        duplicate_df = find_multi_account_customers(vpp_df)

        # Export results
        print("Exporting results...")
        file_path = export_to_excel(summary_df, matches_df, duplicate_df)

        print(f"Successfully exported report to: {file_path}")
        tech.send_notification(f"Report completed: {file_path}")

    except Exception as e:
        error_msg = f"Error during execution: {e}"
        print(error_msg)
        tech.send_notification(error_msg)
        raise

    finally:
        if "driver" in locals():
            driver.quit()


if __name__ == "__main__":
    main()
    sys.exit(0)
